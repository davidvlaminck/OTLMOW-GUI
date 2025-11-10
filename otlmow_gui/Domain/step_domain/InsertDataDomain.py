from pathlib import Path
from typing import List, Iterable, Optional, cast, Union

from openpyxl.reader.excel import load_workbook
from otlmow_converter.Exceptions.ExceptionsGroup import ExceptionsGroup

from otlmow_model.OtlmowModel.BaseClasses.OTLObject import OTLObject, \
    dynamic_create_instance_from_uri
from otlmow_model.OtlmowModel.BaseClasses.RelationInteractor import RelationInteractor
from otlmow_model.OtlmowModel.Classes.Agent import Agent
from otlmow_model.OtlmowModel.Classes.ImplementatieElement.AIMObject import AIMObject
from otlmow_model.OtlmowModel.Classes.ImplementatieElement.RelatieObject import RelatieObject
from otlmow_model.OtlmowModel.Helpers import OTLObjectHelper, RelationValidator

from otlmow_gui.Domain import global_vars
from otlmow_gui.Domain.project.ProjectFile import ProjectFile
from otlmow_gui.Domain.util.CombineAssetHelper import CombineAssetHelper
from otlmow_gui.Domain.util.Helpers import Helpers
from otlmow_gui.Domain.util.SDFHandler import SDFHandler
from otlmow_gui.Domain.logger.OTLLogger import OTLLogger
from otlmow_gui.Domain.project.Project import Project
from otlmow_gui.Domain.step_domain.RelationChangeDomain import RelationChangeDomain, \
    async_save_assets
from otlmow_gui.Domain.enums import FileState
from otlmow_gui.Exceptions.NoIdentificatorError import NoIdentificatorError
from otlmow_gui.Exceptions.RelationHasInvalidTypeUriForSourceAndTarget import \
    RelationHasInvalidTypeUriForSourceAndTarget
from otlmow_gui.Exceptions.RelationHasNonExistingTypeUriForSourceOrTarget import \
    RelationHasNonExistingTypeUriForSourceOrTarget
from otlmow_gui.GUI.dialog_windows.LoadingImageWindow import add_loading_screen_no_delay
from otlmow_gui.GUI.screens.RelationChange_elements.RelationChangeHelpers import RelationChangeHelpers
from otlmow_gui.GUI.translation.GlobalTranslate import GlobalTranslate




class InsertDataDomain:
    """
    Handles operations related to inserting and validating documents wih objects.

    This class provides methods for managing the insertion of data, including
    loading and validating documents, managing relations, and synchronizing
    data between the backend and frontend. It also includes functionality for
    handling exceptions related to data integrity.

    :param None:
    :attributes:
        project (optional): The current project associated with the data insertion.
        collector (OSLOCollector): Collector for managing OSLO data.
        selected_object (optional): The currently selected object in the UI.
        shown_objects (list): List of objects currently displayed in the UI.
        internal_objects (list): List of internal objects related to the project.
        external_objects (list): List of external objects related to the project.
        existing_relations (list): List of existing relations between objects.
        possible_relations_per_class_dict (dict): Dictionary of possible relations per class.
        possible_object_to_object_relations_dict (dict): Dictionary of possible relations between objects.
        aim_id_relations (list): List of relations identified by AIM IDs.
        no_id_count (int): Counter for objects without identifiers.
        external_object_added (bool): Flag indicating if an external object has been added.
    """

    @classmethod
    def init_static(cls,force_refresh = False):
        """
        Initializes static resources for the InsertDataDomain class.

        This method synchronizes the backend documents with the frontend, ensuring that
        the data displayed to the user is up-to-date and consistent with the backend state.

        :param cls: The class itself.
        :returns: None
        """
        if not Helpers.all_OTL_asset_types_dict or force_refresh:
            Helpers.create_external_typeURI_options()

        cls.get_screen().clear_feedback()
        cls.update_frontend()

    @classmethod
    async def check_document(cls, doc_location: Union[str, Path], delimiter: str=";") -> Iterable[OTLObject]:
        """
        Checks a document and converts it into a list of OTL objects.

        This method takes the location of a document and uses the OtlmowConverter
        to convert the contents of the file into a collection of OTL objects,
        including tab information for each object.

        :param cls: The class itself.
        :param doc_location: The path to the document to be checked.
        :type doc_location: str
        :returns: An iterable of OTL objects extracted from the document.
        """
        if isinstance(doc_location,str):
            doc_location_path = Path(doc_location)
        else:
            doc_location_path = doc_location

        exception_group = None
        try:
            if doc_location_path.suffix in ['.xls', '.xlsx']:
                temp_path = InsertDataDomain.remove_dropdown_values_from_excel(doc=doc_location_path)
                assets, exception_group =  await Helpers.converter_from_file_to_object_async(
                    file_path=temp_path,include_tab_info=True)

            elif doc_location_path.suffix == '.sdf':
                # SDF files will make multiple CSV files, one for each class
                temp_path_list = InsertDataDomain.create_temporary_SDF_conversion_to_CSV_files(
                    sdf_filepath=doc_location_path)

                assets = []
                sdf_exception_list = []
                for temp_path in temp_path_list:
                    assets_subset, exception_group_subset = await Helpers.converter_from_file_to_object_async(
                        file_path=Path(temp_path),
                        delimiter=",",
                        include_tab_info=True,
                        allow_non_otl_conform_attributes=True)

                    for asset in assets_subset:
                        if hasattr(asset,"FeatId"):
                            delattr(asset,"FeatId")

                    assets.extend(assets_subset)
                    if exception_group is not None:
                        sdf_exception_list.extend(exception_group.exceptions)

                exception_group = ExceptionsGroup(
                    message=f'Failed to create objects from Excel file {doc_location_path}')
                for exception in sdf_exception_list:
                    exception_group.add_exception(exception)
            else:
                assets, exception_group = await Helpers.converter_from_file_to_object_async(
                    file_path=doc_location_path,include_tab_info=True )

            # second checks done by the GUI
            if exception_group is None:
                exception_group = ExceptionsGroup(
                    message=f'Failed to create objects from Excel file {doc_location_path}')

            cls.check_for_empty_identificators(assets=assets, exception_group=exception_group)

        except ExceptionsGroup as group:
            exception_group = group
            assets = group.objects

        return assets,exception_group

    @classmethod
    def remove_dropdown_values_from_excel(cls, doc: Path) -> Path:
        """
        Removes specific dropdown sheets from an Excel document.

        This method loads an Excel workbook from the specified document path,
        removes the sheets named 'Keuzelijsten' and 'dropdownvalues' if they exist,
        and saves the modified workbook to a temporary path. It returns the path
        of the modified workbook.

        :param cls: The class itself.
        :param doc: The path to the Excel document to be modified.
        :type doc: Path
        :returns: The path to the modified Excel document.
        """

        OTLLogger.logger.debug("starting excel changes")
        wb = load_workbook(doc)
        temp_path = Helpers.create_temp_path(path_to_template_file_and_extension=doc)
        if 'Keuzelijsten' in wb.sheetnames:
            wb.remove(wb['Keuzelijsten'])
        if 'dropdownvalues' in wb.sheetnames:
            wb.remove(wb['dropdownvalues'])

        wb.save(temp_path)
        return temp_path

    @classmethod
    def create_temporary_SDF_conversion_to_CSV_files(cls, sdf_filepath: Path) -> list[str]:

        temp_csv_path = Helpers.create_temp_path(path_to_template_file_and_extension=sdf_filepath).with_suffix(".csv")
        temp_csv_path_str_list = SDFHandler.convert_SDF_to_CSV(sdf_filepath=sdf_filepath,csv_output_path=temp_csv_path)

        return temp_csv_path_str_list


    @classmethod
    def add_template_file_to_project(cls, filepath: Path, project: Project,
                                     state: FileState) -> None:
        """Adds a template file to the specified project.

        This method checks the file extension of the provided template file and, if it is an Excel file,
        removes any dropdown values before adding it to the project. It then copies the file to the project
        and synchronizes the backend documents with the frontend.

        :param cls: The class itself.
        :param filepath: The path to the template file to be added.
        :type filepath: Path
        :param project: The project to which the template file will be added.
        :type project: Project
        :param state: The state of the file being added, indicating its status.
        :type state: FileState
        :returns: None
        """

        if Path(filepath).suffix in ['.xls', '.xlsx']:
            filepath = cls.remove_dropdown_values_from_excel(doc=filepath)

        project.copy_and_add_project_file(file_path=filepath, state=state)
        cls.update_frontend()

    @classmethod
    def return_temporary_path(cls, file_path: Path) -> Path:
        """
        Returns a temporary path based on the file type of the provided file path.

        This method checks the file extension of the given file path and returns a temporary path
        accordingly. If the file is an Excel file, it removes dropdown values; if it is a CSV file,
        it creates a temporary path for the file.

        :param cls: The class itself.
        :param file_path: The path to the file for which a temporary path is requested.
        :type file_path: Path
        :returns: The temporary path for the processed file.
        :rtype: Path
        """

        if Path(file_path).suffix in ['.xls', '.xlsx']:
            return cls.remove_dropdown_values_from_excel(doc=file_path)
        elif Path(file_path).suffix == '.csv':
            return Helpers.create_temp_path(path_to_template_file_and_extension=file_path)

    @classmethod
    def add_files_to_backend_list(cls, files: list[str],
                                  states: Optional[list[FileState]] = None) -> None:
        """
        Adds a list of files to the backend project with specified states.

        This method takes a list of file paths and an optional list of states,
        adding each file to the current project in the backend. If no states are
        provided, it defaults to a warning state for each file, and updates the
        file list in the user interface after the operation.

        :param cls: The class itself.
        :param files: A list of file paths to be added to the backend project.
        :type files: list[str]
        :param states: An optional list of states corresponding to each file.
        :type states: Optional[list[FileState]]
        :returns: None
        """

        if states is None:
            states = [FileState.WARNING for _ in range(len(files))]

        for i in range(len(files)):
            InsertDataDomain.add_template_file_to_project(project=global_vars.current_project,
                                                          filepath=Path(files[i]),
                                                          state=states[i])

        cls.get_screen().update_file_list()

    @classmethod
    def update_frontend(cls) -> bool:
        """
        Synchronizes the backend project files with the frontend display.

        This method clears the current list of project files in the frontend and
        populates it with the saved project files from the backend. It checks the
        state of each file and returns a boolean indicating whether all files are valid.

        :param cls: The class itself.
        :returns: True if all files are valid; otherwise, False.
        :rtype: bool
        """
        cls.get_screen().project_files_overview_field.clear()
        cls.get_screen().add_file_overview_placeholder_to_front_end_list()
        files_in_project = global_vars.current_project.get_saved_projectfiles()
        all_valid = False
        if files_in_project:
            cls.get_screen().project_files_overview_field.clear()

            all_valid = True
            for item in global_vars.current_project.get_saved_projectfiles():
                cls.get_screen().add_file_to_frontend_list(file=item.file_path,asset_state=item.state)
                if item.state != FileState.OK:
                    all_valid = False

        cls.get_screen().update_control_button_state()
        global_vars.otl_wizard.main_window.enable_steps()
        return all_valid

    @classmethod
    def delete_backend_document(cls, item_file_path: str) -> None:
        """Deletes a document from the backend project.

        This method removes the specified file from the current project's list of
        project files and synchronizes the backend documents with the frontend
        display to reflect the changes.

        :param cls: The class itself.
        :param item_file_path: The path of the file to be deleted from the backend.
        :type item_file_path: str
        :returns: None

        """
        global_vars.current_project.remove_project_file(Path(item_file_path))
        cls.get_screen().clear_feedback()

        InsertDataDomain.update_frontend()

    @classmethod
    def check_current_project_project_files_existence(cls) -> list[ProjectFile]:
        missing_project_files = global_vars.current_project.check_if_project_files_exist()

        if len(missing_project_files):
            global_vars.current_project.save_project_filepaths_to_file()
            cls.update_frontend()

        return missing_project_files
    @classmethod
    @add_loading_screen_no_delay
    @async_save_assets
    async def load_and_validate_documents(cls,**kwargs) -> tuple[list[dict], list]:
        """
        Loads and validates documents from the current project's saved files.

        This method iterates through the saved project files, attempting to load
        and validate each document. It collects any errors encountered during the
        process and returns a list of errors along with a list of successfully
        loaded objects.

        :param cls: The class itself.
        :returns: A tuple containing:
            - A list of error dictionaries, each containing an exception and the
              associated file path.
            - A list of successfully loaded objects.
        :rtype: tuple[list[dict],list]
        """
        OTLLogger.logger.debug(f"Executing InsertDataDomain.load_and_validate_documents() for project {global_vars.current_project.eigen_referentie}", extra={"timing_ref": f"validate_{global_vars.current_project.eigen_referentie}"})
        error_set: list[dict] = []
        relations_per_filepath_str_dict: dict[str,list[RelatieObject]] = {}
        assets_per_filepath_str_dict: dict[str, list[AIMObject]] = {}

        for project_file in global_vars.current_project.get_saved_projectfiles():
            file_path = project_file.file_path
            try:

                objects, exception_group = await cls.check_document(file_path)

                assets, relations = cls.divide_otl_objects(objects)

                if len(relations):
                    relations_per_filepath_str_dict[str(file_path)] = relations
                if len(assets):
                    # noinspection PyTypeChecker
                    assets_per_filepath_str_dict[str(file_path)] = assets

                if len(exception_group.exceptions) > 0:
                    raise exception_group

                project_file.state = FileState.OK

            except Exception as ex:
                error_set.append({"exception": ex, "path_str": file_path})
                project_file.state = FileState.ERROR

        total_assets = cls.flatten_list(objects_lists=list(assets_per_filepath_str_dict.values()))

        # check if there are relations without a bron/doel.TypeURI that cannot be tied to an asset
        for project_file in global_vars.current_project.get_saved_projectfiles():
            file_path = project_file.file_path
            file_path_str = str(file_path)

            if file_path_str not in relations_per_filepath_str_dict.keys():
                continue

            try:
                exception_group = ExceptionsGroup(
                    message=f'Failed to create objects from Excel file {file_path_str}')

                cls.check_for_invalid_relations(relations=relations_per_filepath_str_dict[file_path_str],
                                                ref_assets=total_assets, exception_group=exception_group)

                if len(exception_group.exceptions) > 0:
                    relations_per_filepath_str_dict.pop(file_path_str)
                    raise exception_group

                if project_file.state != FileState.ERROR:
                    project_file.state = FileState.OK
            except Exception as ex:
                error_set.append({"exception": ex, "path_str": file_path})
                project_file.state = FileState.ERROR



        if len(error_set):
            OTLLogger.logger.debug(
                f"Executing InsertDataDomain.load_and_validate_documents() for project {global_vars.current_project.eigen_referentie} (INVALID)",
                extra={"timing_ref": f"validate_{global_vars.current_project.eigen_referentie}"})
            # state can be changed to either OK or ERROR
            global_vars.current_project.save_project_filepaths_to_file()
            cls.update_frontend()
            return error_set, []

        objects_in_memory = cls.flatten_list(objects_lists=list(assets_per_filepath_str_dict.values()))
        # noinspection PyTypeChecker
        objects_in_memory.extend(cls.flatten_list(objects_lists=list(relations_per_filepath_str_dict.values())))

        try:
            objects_in_memory = await cls.combine_assets_wrapper(objects_in_memory)
        except ExceptionsGroup as ex:
            error_set.append({"exception": ex, "path_str": GlobalTranslate._("[FILE INFO NOT AVAILABLE]")})
            # the be safe set all document to invalid
            for project_file in global_vars.current_project.get_saved_projectfiles():
                project_file.state = FileState.ERROR

        # state can be changed to either OK or ERROR
        global_vars.current_project.save_project_filepaths_to_file()
        cls.update_frontend()

        # passing objects to other screens
        # global_vars.otl_wizard.main_window.step3_visuals.create_html(
        #     objects_in_memory=objects_in_memory)
        await RelationChangeDomain.set_instances(objects_list=objects_in_memory)
        # global_vars.otl_wizard.main_window.step3_visuals.reload_html()
        global_vars.current_project.visualisation_uptodate.set_clear_all(True)
        global_vars.current_project.save_visualisation_python_support_data()

        object_count = len(objects_in_memory)
        OTLLogger.logger.debug(
            f"Executing InsertDataDomain.load_and_validate_documents() for project {global_vars.current_project.eigen_referentie} ({object_count} objects)",
            extra={"timing_ref": f"validate_{global_vars.current_project.eigen_referentie}"})

        return error_set, objects_in_memory

    @classmethod
    async def combine_assets_wrapper(cls, objects_in_memory):

        objects_in_memory = CombineAssetHelper.combine_assets(objects_in_memory)

        return objects_in_memory

    @classmethod
    def divide_otl_objects(cls, objects):
        relations: list[RelatieObject] = []
        assets: list[OTLObject] = []
        for otl_object in objects:
            if OTLObjectHelper.is_relation(otl_object=otl_object):
                relations.append(otl_object)
            else:
                assets.append(otl_object)
        return assets, relations

    @classmethod
    def check_for_invalid_relations(cls, ref_assets: list[AIMObject], relations: list[RelatieObject],
                                    exception_group: ExceptionsGroup) -> None:
        """
        Checks for invalid relations among the provided assets.

        This method iterates through a list of assets and verifies the validity of
        relations defined within them. It adds any detected issues to the provided
        exception group, ensuring that all invalid relations are reported for further
        handling.

        :param cls: The class itself.
        :param ref_assets: A list of OTL objects to be checked for invalid relations.
        :type ref_assets: list[OTLObject]
        :param exception_group: A group to collect exceptions related to invalid relations.
        :type exception_group: ExceptionsGroup
        :returns: None
        """
        id_to_typeURI_dict: Optional[dict[str,str]] = None
        for relation in relations:

            bron_type_uri = Helpers.extract_corrected_relation_partner_typeURI(
                relation.bron.typeURI,
                relation.bronAssetId.identificator,
                id_to_typeURI_dict,
                ref_assets)

            if not bron_type_uri:
                ex = RelationHasNonExistingTypeUriForSourceOrTarget(
                    relation_type_uri=relation.typeURI,
                    relation_identificator=relation.assetId.identificator,
                    wrong_field="bron.typeURI",
                    wrong_value=None,
                    tab=RelationChangeHelpers.get_abbreviated_typeURI(
                        relation.typeURI, False))
                exception_group.add_exception(error=ex)

            doel_type_uri = Helpers.extract_corrected_relation_partner_typeURI(
                relation.doel.typeURI,
                relation.doelAssetId.identificator,
                id_to_typeURI_dict,
                ref_assets)

            if not doel_type_uri:
                ex = RelationHasNonExistingTypeUriForSourceOrTarget(
                    relation_type_uri=relation.typeURI,
                    relation_identificator=relation.assetId.identificator,
                    wrong_field="bron.typeURI",
                    wrong_value=None,
                    tab=RelationChangeHelpers.get_abbreviated_typeURI(
                        relation.typeURI, False))
                exception_group.add_exception(error=ex)

            if not bron_type_uri or not doel_type_uri:
                # the rest of the check will fail anyway
                continue

            has_valid_bron_typeURI:bool = InsertDataDomain.does_typeURI_exist(bron_type_uri)
            if not has_valid_bron_typeURI:
                ex = RelationHasNonExistingTypeUriForSourceOrTarget(
                    relation_type_uri=relation.typeURI,
                    relation_identificator=relation.assetId.identificator,
                    wrong_field="bron.typeURI",
                    wrong_value=bron_type_uri,
                    tab=RelationChangeHelpers.get_abbreviated_typeURI(
                        relation.typeURI,False))
                exception_group.add_exception(error=ex)

            has_valid_doel_typeURI: bool = InsertDataDomain.does_typeURI_exist(doel_type_uri)
            if not has_valid_doel_typeURI:
                ex = RelationHasNonExistingTypeUriForSourceOrTarget(
                    relation_type_uri=relation.typeURI,
                    relation_identificator=relation.assetId.identificator,
                    wrong_field="doel.typeURI",
                    wrong_value=doel_type_uri,
                    tab=RelationChangeHelpers.get_abbreviated_typeURI(
                        relation.typeURI,False))
                exception_group.add_exception(error=ex)


            if not has_valid_bron_typeURI or not has_valid_doel_typeURI:
                # the rest of the check will fail anyway
                continue

            try:
                is_valid_relation = RelationValidator.is_valid_relation(
                    relation_type=type(relation),
                    source_typeURI=bron_type_uri,
                    target_typeURI=doel_type_uri)
            except Exception as e:
                OTLLogger.logger.warning(e)
                is_valid_relation = False

            if not is_valid_relation:
                ex = cls.raise_wrong_doel_or_target(relation=relation,
                                                    tab=RelationChangeHelpers.get_abbreviated_typeURI(
                                                        relation.typeURI, False),
                                                    bron_type_uri=bron_type_uri,
                                                    doel_type_uri=doel_type_uri)
                exception_group.add_exception(error=ex)

    @classmethod
    def does_typeURI_exist(cls, bron_type_uri):
        return bron_type_uri in Helpers.all_OTL_asset_types_dict.values()



    @classmethod
    def raise_wrong_doel_or_target(cls, relation: RelatieObject, tab: str, bron_type_uri,
                                   doel_type_uri) -> RelationHasInvalidTypeUriForSourceAndTarget:
        """
        Creates an exception for an invalid relation with incorrect source or target type URIs.

        This method constructs and returns an instance of 
        RelationHasInvalidTypeUriForSourceAndTarget, populated with details about 
        the invalid relation, including the type URIs and the associated tab.

        :param doel_type_uri:
        :param bron_type_uri:
        :param cls: The class itself.
        :param relation: The relation object that contains the type URIs to be validated.
        :type relation: RelatieObject
        :param tab: The name of the tab where the relation is defined.
        :type tab: str
        :returns: An instance of RelationHasInvalidTypeUriForSourceAndTarget.
        :rtype: RelationHasInvalidTypeUriForSourceAndTarget
        """

        return RelationHasInvalidTypeUriForSourceAndTarget(
            relation_type_uri=relation.typeURI,
            relation_identificator=relation.assetId.identificator,
            wrong_field="bron.typeURI",
            wrong_value=bron_type_uri,
            wrong_field2="doel.typeURI",
            wrong_value2=doel_type_uri,
            tab=tab)

    @classmethod
    def flatten_list(cls, objects_lists: list[list[AIMObject]]) -> list[AIMObject]:
        """Flattens a list of lists into a single list of AIM objects.

        This method combines multiple lists of AIM objects into a single list, 
        making it easier to process all objects at once. It iterates through each 
        sublist and extends the main list with the objects contained in them.

        :param cls: The class itself.
        :param objects_lists: A list of lists containing AIM objects to be flattened.
        :type objects_lists: list[list[AIMObject]]
        :returns: A single list containing all AIM objects from the input lists.
        :rtype: list[AIMObject]
        """

        objects_in_memory: List[AIMObject] = []
        for objects_list in objects_lists:
            objects_in_memory.extend(objects_list)
        return objects_in_memory

    @classmethod
    def get_screen(cls):
        return global_vars.otl_wizard.main_window.step2

    @classmethod
    def detect_more_complex_target_or_source_typeURI_errors(cls, relation: RelatieObject) -> None:
        """
        @Unused
        Detects complex errors related to source or target type URIs in a relation.

        This method validates the relationship between the source and target assets, 
        checking for the existence of valid relations. If the relations are invalid, 
        it raises an appropriate error indicating the specific issue with the source 
        or target type URIs.

        :param cls: The class itself.
        :param relation: The relation object to be validated.
        :returns: None
        """

        if RelationValidator.is_valid_relation(relation_type=type(relation),
                                                   source_typeURI=relation.bron.typeURI,
                                                   target_typeURI=relation.doel.typeURI):
            return

        # RelationValidator.is_valid_relation doesn't say if bron or doel is wrong
        source_instance:RelationInteractor = dynamic_create_instance_from_uri(relation.bron.typeURI)
        concrete_source_relations = list(source_instance._get_all_concrete_relations())
        if concrete_source_relations_of_type_relation := {
            rel for rel in concrete_source_relations if rel[1] == relation.typeURI
        }:
            if  [
                rel
                for rel in concrete_source_relations_of_type_relation
                if rel[2] == relation.doel.typeURI]:
                OTLLogger.logger.debug("Error in logic")
            else:
                # source asset doesn't have this relation to target
                cls.raise_wrong_doel_or_target(relation=relation,
                                               tab=RelationChangeHelpers.get_abbreviated_typeURI(
                                                   typeURI=relation.typeURI,
                                                   add_namespace=False),
                                               bron_type_uri=relation.bron.typeURI,
                                               doel_type_uri=relation.doel.typeURI)
        else:
            # target asset has relation but not to source
            cls.raise_wrong_doel_or_target(relation=relation,
                                           tab=RelationChangeHelpers.get_abbreviated_typeURI(
                                               typeURI=relation.typeURI,
                                               add_namespace=False),
                                           bron_type_uri=relation.bron.typeURI,
                                           doel_type_uri=relation.doel.typeURI)

    @classmethod
    def check_for_empty_identificators(cls, assets: Iterable[OTLObject],
                                       exception_group: ExceptionsGroup) -> None:
        """Checks for empty identificators in the provided assets.

        This method iterates through a collection of assets and verifies that each asset 
        has a valid identificator. If an asset is found to be missing an identificator, 
        it adds an exception to the provided exception group.

        :param cls: The class itself.
        :param assets: An iterable collection of OTL objects to be checked for identificators.
        :type assets: Iterable[OTLObject]
        :param exception_group: A group to collect exceptions related to missing identificators.
        :type exception_group: ExceptionsGroup
        :returns: None
        """

        for asset in assets:
            identificator = None
            if asset.typeURI == 'http://purl.org/dc/terms/Agent':
                new_external_agent: Agent = cast(Agent, asset)
                identificator = new_external_agent.agentId.identificator

            else:
                new_external_asset: AIMObject = cast(AIMObject, asset)
                identificator = new_external_asset.assetId.identificator

            if not identificator:
                exception_group.add_exception(error=NoIdentificatorError(GlobalTranslate._ ,original_exception=ValueError(),tab=RelationChangeHelpers.get_abbreviated_typeURI(asset.typeURI,False)))
