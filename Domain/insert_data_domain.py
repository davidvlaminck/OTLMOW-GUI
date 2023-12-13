import json
import logging
import os
import tempfile
from enum import Enum
from pathlib import Path
from typing import List

from openpyxl.reader.excel import load_workbook
from otlmow_converter.OtlmowConverter import OtlmowConverter

from Domain import global_vars
from Domain.Project import Project
from Domain.ProjectFileManager import ProjectFileManager
from Domain.project_file import ProjectFile


class InsertDataDomain:

    @classmethod
    def check_document(cls, doc_location) -> List:
        converter = OtlmowConverter()
        return converter.create_assets_from_file(filepath=Path(doc_location))

    @classmethod
    def start_excel_changes(cls, doc) -> Path:
        wb = load_workbook(doc)
        temp_path = cls.create_temp_path(path_to_template_file_and_extension=doc)
        if 'Keuzelijsten' in wb.sheetnames:
            wb.remove(wb['Keuzelijsten'])
        wb.save(temp_path)
        return temp_path

    @classmethod
    def create_temp_path(cls, path_to_template_file_and_extension: Path) -> Path:
        tempdir = Path(tempfile.gettempdir()) / 'temp-otlmow'
        if not tempdir.exists():
            os.makedirs(tempdir)
        doc_name = Path(path_to_template_file_and_extension).name
        return Path(tempdir) / doc_name

    @classmethod
    def add_template_file_to_project(cls, filepath: Path, project: Project, state: Enum):
        if Path(filepath).suffix in ['.xls', '.xlsx']:
            filepath = cls.start_excel_changes(doc=filepath)
        end_loc = ProjectFileManager().add_template_file_to_project(filepath=filepath)
        template_file = ProjectFile(file_path=end_loc, state=state)
        project.templates_in_memory.append(template_file)

    @classmethod
    def return_temporary_path(cls, file_path: Path) -> Path:
        if Path(file_path).suffix in ['.xls', '.xlsx']:
            temp_path = cls.start_excel_changes(doc=file_path)
        elif Path(file_path).suffix == '.csv':
            temp_path = cls.create_temp_path(path_to_template_file_and_extension=file_path)
        return temp_path

    @classmethod
    def delete_project_file_from_project(cls, project: Project, file_path: Path):
        for file in project.templates_in_memory:
            if file.file_path == file_path:
                project.templates_in_memory.remove(file)
                break
        ProjectFileManager().delete_template_file_from_project(file_path=file_path)
        ProjectFileManager().add_project_files_to_file(project=project)

    @classmethod
    def remove_all_project_files(cls, project):
        project_file_manager = ProjectFileManager()
        logging.debug("memory contains %s", project.templates_in_memory)
        for file in project.templates_in_memory:
            logging.debug("starting to delete file %s", file.file_path)
            project_file_manager.delete_template_file_from_project(file_path=file.file_path)
        project.templates_in_memory = []
        ProjectFileManager().add_project_files_to_file(project=project)
