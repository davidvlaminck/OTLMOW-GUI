import logging
import os
import tempfile

from pathlib import Path
from typing import List, Iterable, Optional, cast

from openpyxl.reader.excel import load_workbook
from otlmow_converter.OtlmowConverter import OtlmowConverter
from otlmow_model.OtlmowModel.BaseClasses.OTLObject import OTLObject, \
    dynamic_create_instance_from_uri
from otlmow_model.OtlmowModel.Classes.ImplementatieElement.RelatieObject import RelatieObject
from otlmow_model.OtlmowModel.Helpers import OTLObjectHelper, RelationValidator

from Domain import global_vars
from Domain.Project import Project
from Domain.RelationChangeDomain import RelationChangeDomain, save_assets
from Domain.enums import FileState
from Exceptions.RelationHasInvalidTypeUriForSourceAndTarget import \
    RelationHasInvalidTypeUriForSourceAndTarget
from Exceptions.RelationHasNonExistingTypeUriForSourceOrTarget import \
    RelationHasNonExistingTypeUriForSourceOrTarget
from UnitTests.TestClasses.Classes.ImplementatieElement.AIMObject import AIMObject


class InsertDataDomain:

    @classmethod
    def init_static(cls):
        cls.sync_backend_documents_with_frontend()

    @classmethod
    def check_document(cls, doc_location) -> Iterable[OTLObject]:
        return OtlmowConverter.from_file_to_objects(file_path=Path(doc_location))

    @classmethod
    def remove_dropdown_values_from_excel(cls, doc) -> Path:
        logging.debug("starting excel changes")
        wb = load_workbook(doc)
        temp_path = cls.create_temp_path(path_to_template_file_and_extension=doc)
        if 'Keuzelijsten' in wb.sheetnames:
            wb.remove(wb['Keuzelijsten'])
        if 'dropdownvalues' in wb.sheetnames:
            wb.remove(wb['dropdownvalues'])

        wb.save(temp_path)
        return temp_path

    @classmethod
    def create_temp_path(cls, path_to_template_file_and_extension: Path) -> Path:
        tempdir = Path(tempfile.gettempdir()) / 'temp-otlmow'
        if not tempdir.exists():
            os.makedirs(tempdir)
        doc_name = Path(path_to_template_file_and_extension).name
        return Path(tempdir) / doc_name

    @classmethod
    def add_template_file_to_project(cls, filepath: Path, project: Project, state: FileState):
        if Path(filepath).suffix in ['.xls', '.xlsx']:
            filepath = cls.remove_dropdown_values_from_excel(doc=filepath)

        end_loc = project.make_copy_of_added_file(filepath=filepath)
        project.add_saved_project_file(file_path=end_loc, state=state)
        cls.sync_backend_documents_with_frontend()


    @classmethod
    def return_temporary_path(cls, file_path: Path) -> Path:
        if Path(file_path).suffix in ['.xls', '.xlsx']:
            return cls.remove_dropdown_values_from_excel(doc=file_path)
        elif Path(file_path).suffix == '.csv':
            return cls.create_temp_path(path_to_template_file_and_extension=file_path)

    @classmethod
    def add_files_to_backend_list(cls, files: list[str], states: Optional[list[FileState]] = None):
        if states is None:
            states = [FileState.WARNING for _ in range(len(files))]

        for i in range(len(files)):
            InsertDataDomain.add_template_file_to_project(project=global_vars.current_project,
                                                          filepath=Path(files[i]),
                                                          state=states[i])

        cls.get_screen().update_file_list()

    @classmethod
    def sync_backend_documents_with_frontend(cls) -> bool:
        all_valid = True
        cls.get_screen().project_files_overview_field.clear()
        for item in global_vars.current_project.get_saved_projectfiles():
            cls.get_screen().add_file_to_frontend_list(item.file_path,item.state)
            if item.state != FileState.OK:
                all_valid = False

        return all_valid

    @classmethod
    def delete_backend_document(cls, item_file_path: str):
        global_vars.current_project.remove_project_file(Path(item_file_path))

        InsertDataDomain.sync_backend_documents_with_frontend()

    @classmethod
    @save_assets
    def load_and_validate_documents(cls):
        error_set: list[dict] = []
        objects_lists = []

        for project_file in global_vars.current_project.get_saved_projectfiles():


            try:

                file_path = project_file.file_path
                temp_path = ""
                if file_path.suffix in ['.xls', '.xlsx']:
                    temp_path = InsertDataDomain.remove_dropdown_values_from_excel(doc=file_path)
                elif file_path.suffix == '.csv':
                    temp_path = file_path

                assets = InsertDataDomain.check_document(doc_location=temp_path)

                cls.check_for_invalid_relations(assets)

                project_file.state = FileState.OK
                objects_lists.append(assets)
            except Exception as ex:
                error_set.append({"exception": ex, "path_str": file_path})
                # ProjectFileManager.add_template_file_to_project(project=global_vars.current_project,
                #                                               filepath=Path(doc),
                #                                               state=FileState.ERROR)
                project_file.state = FileState.ERROR

        # state can be changed to either OK or ERROR
        cls.sync_backend_documents_with_frontend()

        objects_in_memory = cls.flatten_list(objects_lists)

        global_vars.otl_wizard.main_window.step3_visuals.create_html(objects_in_memory)
        RelationChangeDomain.set_instances(objects_in_memory)

        return error_set, objects_lists

    @classmethod
    def check_for_invalid_relations(cls, assets):
        for asset in assets:
            if OTLObjectHelper.is_relation(asset):
                relation = cast(RelatieObject, asset)
                if relation.bron.typeURI not in RelationChangeDomain.all_OTL_asset_types_dict.values():
                    raise RelationHasNonExistingTypeUriForSourceOrTarget(relation.typeURI,
                                                                         relation.assetId.identificator,
                                                                         "bron.typeURI",
                                                                         relation.bron.typeURI)
                if relation.doel.typeURI not in RelationChangeDomain.all_OTL_asset_types_dict.values():
                    raise RelationHasNonExistingTypeUriForSourceOrTarget(relation.typeURI,
                                                                         relation.assetId.identificator,
                                                                         "doel.typeURI",
                                                                         relation.doel.typeURI)

                # cls.detect_more_complex_target_or_source_typeURI_errors(relation)
                if not RelationValidator.is_valid_relation(relation_type=type(relation),
                                                             source_typeURI=relation.bron.typeURI,
                                                             target_typeURI=relation.doel.typeURI):
                    cls.raise_wrong_doel_or_target(relation)



    @classmethod
    def raise_wrong_doel_or_target(cls, relation):
        raise RelationHasInvalidTypeUriForSourceAndTarget(relation.typeURI,
                                                          relation.assetId.identificator,
                                                          "bron.typeURI",
                                                          relation.bron.typeURI,
                                                          "doel.typeURI",
                                                          relation.doel.typeURI)

    @classmethod
    def flatten_list(cls, objects_lists):
        objects_in_memory: List[AIMObject] = []
        for objects_list in objects_lists:
            objects_in_memory.extend(objects_list)
        return objects_in_memory

    @classmethod
    def get_screen(cls):
        return global_vars.otl_wizard.main_window.step2

    @classmethod
    def detect_more_complex_target_or_source_typeURI_errors(cls, relation):
        if not RelationValidator.is_valid_relation(relation_type=type(relation),
                                                   source_typeURI=relation.bron.typeURI,
                                                   target_typeURI=relation.doel.typeURI):

            # RelationValidator.is_valid_relation doesn't say if bron or doel is wrong
            source_instance = dynamic_create_instance_from_uri(relation.bron.typeURI)
            concrete_source_relations = list(source_instance._get_all_concrete_relations())
            concrete_source_relations_of_type_relation = set(
                [rel for rel in concrete_source_relations if rel[1] == relation.typeURI])

            if concrete_source_relations_of_type_relation:
                # source asset has relation
                concrete_source_relation_to_target = \
                    [rel for rel in concrete_source_relations_of_type_relation if
                     rel[2] == relation.doel.typeURI]
                if not concrete_source_relation_to_target:
                    # source asset doesn't have this relation to target
                    cls.raise_wrong_doel_or_target(relation)
                else:
                    logging.debug("Error in logic")
            else:
                target_instance = dynamic_create_instance_from_uri(
                    relation.doel.typeURI)
                concrete_target_relations = list(
                    target_instance._get_all_concrete_relations())
                concrete_target_relations_of_type_relation = set(
                    [rel for rel in concrete_target_relations if
                     rel[1] == relation.typeURI])
                if concrete_target_relations_of_type_relation:
                    # target asset has relation but not to source
                    cls.raise_wrong_doel_or_target(relation)
                else:
                    # both target and source asset do not have relation
                    cls.raise_wrong_doel_or_target(relation)

