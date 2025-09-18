import asyncio
import shutil
from pathlib import Path

import openpyxl
import time
import csv
import tempfile
import pytest
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QLineEdit, QDialogButtonBox, QApplication, QPushButton

import otlmow_gui.Domain.global_vars as global_vars
from otlmow_gui.Domain.logger.OTLLogger import OTLLogger
from otlmow_gui.GUI.MainWindow import MainWindow
from otlmow_gui.GUI.dialog_windows.UpsertProjectWindow import UpsertProjectWindow
from otlmow_gui.GUI.dialog_windows.file_picker_dialog.SubsetLoadFilePickerDialog import SubsetLoadFilePickerDialog
from otlmow_gui.GUI.translation.GlobalTranslate import GlobalTranslate
from otlmow_gui.wizard_main import Settings, LANG_DIR

# --- Fixtures and helpers unchanged ---

@pytest.fixture(autouse=True)
def patch_upsert_project_window_exec():
    original_exec = UpsertProjectWindow.exec
    def fake_exec(self):
        self.show()
        return 0
    UpsertProjectWindow.exec = fake_exec
    yield
    UpsertProjectWindow.exec = original_exec


@pytest.fixture
def patch_subset_file_picker(monkeypatch):
    def do_patch(subset_path):
        def fake_summon(self, directory=None):
            return [str(subset_path)]
        monkeypatch.setattr(SubsetLoadFilePickerDialog, "summon", fake_summon)
    return do_patch


@pytest.fixture(autouse=True)
def patch_template_save_file_picker_dialog(monkeypatch):
    from otlmow_gui.GUI.dialog_windows.file_picker_dialog.TemplateSaveFilePickerDialog import TemplateSaveFilePickerDialog
    def fake_summon(self, *args, **kwargs):
        # Try to get the file type from the GUI combobox if possible
        try:
            mw = self.main_window
            filetype = mw.step1.file_extension_selection.currentText()
            project_name = mw.home_screen.table.item(0, 0).text()
        except Exception:
            # fallback to kwargs or defaults
            filetype = kwargs.get("chosen_file_format", "Excel")
            project_name = kwargs.get("project_name", "FullTestRunProject_1b_4")
        ext = ".xlsx" if filetype == "Excel" else (".csv" if filetype == "CSV" else ".xlsx")
        export_path = Path(tempfile.gettempdir()) / f"{project_name}_template{ext}"
        print(f"[DEBUG] TemplateSaveFilePickerDialog.summon called: filetype={filetype}, project_name={project_name}, export_path={export_path}")
        return [export_path]
    monkeypatch.setattr(TemplateSaveFilePickerDialog, "summon", fake_summon)


def wait_for_dialog(qtbot, dialog_cls, timeout=10000):
    def dialog_shown():
        return [w for w in QApplication.instance().allWidgets() if isinstance(w, dialog_cls) and w.isVisible()]
    qtbot.waitUntil(lambda: len(dialog_shown()) > 0, timeout=timeout)
    return dialog_shown()[0]


def fill_upsert_project_dialog(qtbot, dialog, eigen_ref, bestek, subset=None):
    eigen_ref_edit = dialog.findChild(QLineEdit, "eigen_ref_edit")
    assert eigen_ref_edit is not None, "eigen_ref_edit not found"
    eigen_ref_edit.setText(eigen_ref)
    bestek_edit = dialog.findChild(QLineEdit, "bestek_edit")
    assert bestek_edit is not None, "bestek_edit not found"
    bestek_edit.setText(bestek)
    if subset:
        file_picker_btn = dialog.findChild(QPushButton, "file_picker_btn")
        assert file_picker_btn is not None, "file_picker_btn not found"
        qtbot.mouseClick(file_picker_btn, Qt.MouseButton.LeftButton)
        subset_edit = dialog.findChild(QLineEdit, "subset_edit")
        qtbot.waitUntil(lambda: subset_edit.text() == subset, timeout=10000)


def click_ok_and_expect_close(qtbot, dialog):
    button_box = dialog.findChild(QDialogButtonBox)
    ok_button = button_box.button(QDialogButtonBox.StandardButton.Ok)
    assert ok_button is not None, "OK button not found"
    qtbot.mouseClick(ok_button, Qt.MouseButton.LeftButton)
    qtbot.waitUntil(lambda: not dialog.isVisible(), timeout=10000)


# --- Refactored helpers for modular test steps ---

async def create_and_open_project(qtbot, main_window, project_name, bestek, subset_path):
    # Arrange: create new project via dialog
    qtbot.mouseClick(main_window.home_screen.header.new_project_button, Qt.MouseButton.LeftButton)
    dialog = wait_for_dialog(qtbot, UpsertProjectWindow)
    fill_upsert_project_dialog(qtbot, dialog, eigen_ref=project_name, bestek=bestek, subset=str(subset_path))
    click_ok_and_expect_close(qtbot, dialog)
    # Wait for project to appear in table
    qtbot.waitUntil(lambda: main_window.home_screen.table.rowCount() > 0, timeout=10000)
    found = any(main_window.home_screen.table.item(r, 0).text() == project_name
                for r in range(main_window.home_screen.table.rowCount()))
    assert found, f"Project {project_name} not found in table"
    # Open the project using the application logic
    from otlmow_gui.Domain.step_domain.HomeDomain import HomeDomain
    HomeDomain.open_project(project_name)
    # Give the event loop a chance to start async class list fill
    await asyncio.sleep(0.2)
    return main_window.step1.all_classes


def select_classes(class_table, exclude=None):
    # Arrange: select all except those in exclude
    exclude = exclude or set()
    indexes_to_select = [
        class_table.item(r).data(1)[1]
        for r in range(class_table.count())
        if class_table.item(r).text() not in exclude
    ]
    from otlmow_gui.Domain.step_domain.TemplateDomain import TemplateDomain
    TemplateDomain.set_selected_indexes(indexes_to_select)
    class_table.clearSelection()
    for r in range(class_table.count()):
        if class_table.item(r).data(1)[1] in indexes_to_select:
            class_table.item(r).setSelected(True)
    class_table.sync_selected_items_with_backend()


def set_export_options(step1, *, choice_list=True, geometry=True, attr_info=True, example=True, expanded_info=True):
    # Arrange: set export options
    step1.add_choice_list.setChecked(choice_list)
    step1.add_geometry_attributes.setChecked(geometry)
    step1.export_attribute_info.setChecked(attr_info)
    step1.example_amount_checkbox.setChecked(example)
    step1.radio_button_expanded_info.setChecked(expanded_info)


async def export_template_and_wait_for_file(qtbot, export_btn, export_path, timeout=7):
    # Act: click export and wait for file
    if export_path.exists():
        export_path.unlink()
    qtbot.mouseClick(export_btn, Qt.MouseButton.LeftButton)
    await asyncio.sleep(0.2)
    async def wait_for_file(path, poll_interval=0.1):
        while not path.exists():
            await asyncio.sleep(poll_interval)
        return path
    try:
        await asyncio.wait_for(wait_for_file(export_path), timeout=timeout)
    except asyncio.TimeoutError:
        raise TimeoutError(f"File {export_path} was not created after {timeout} seconds")
    assert export_path.exists(), f"File {export_path} was not created"


def verify_exported_classes_in_excel(export_path, expected_classes):
    # Assert: check only expected classes are present in Excel
    wb = openpyxl.load_workbook(export_path)
    expected = set(expected_classes)
    for sheet in wb.sheetnames:
        if sheet == 'Keuzelijsten':
            continue
        ws = wb[sheet]
        uri_value = ws['A'][2].value
        assert uri_value in expected, f"Unexpected value in sheet {sheet}: {uri_value}"
        expected.remove(uri_value)
    assert not expected, f"Not all expected classes were found in the Excel. Missing: {expected}"


def remove_first_row_from_excel(export_path):
    # Act: remove first row from each sheet and save
    wb = openpyxl.load_workbook(export_path)
    for ws in wb.worksheets:
        ws.delete_rows(1)
    wb.save(export_path)

# --- Main test using modular helpers ---

@pytest.mark.gui
@pytest.mark.asyncio
async def test_1_4_template_generation_and_class_selection(qtbot, patch_subset_file_picker):
    """
    TEST stap 1: Aanmaken van template
    - Maak een nieuw project aan (zoals in 1a test_3).
    - Check of alle classes initieel geselecteerd zijn
    - Deselecteer de Contactor en de Kokerafsluiting klasse
    - Genereer een EXCEL file met uitgebreide info en alle andere opties aan
    - Check of alleen de aangeduide 4 klasses aanwezig zijn
    - Verwijder in elke sheet de eerste RIJ met omschrijvingen en save
    """
    project_name = "FullTestRunProject_1b_4"
    project_dir = Path.home() / 'OTLWizardProjects' / 'Projects' / project_name
    export_path = Path(tempfile.gettempdir()) / f"{project_name}_template.xlsx"
    slagboom_subset = Path(__file__).parent.parent.parent / "otlmow_gui" / "demo_projects" / "slagbomen_project" / "voorbeeld-slagboom.db"

    # Clean up before test
    if project_dir.exists():
        shutil.rmtree(project_dir)
    if export_path.exists():
        export_path.unlink()

    try:
        # Arrange: App and main window
        settings = Settings.get_or_create_settings_file()
        OTLLogger.init()
        language = GlobalTranslate(settings, LANG_DIR).get_all()
        main_window = MainWindow(language)
        global_vars.otl_wizard = type("FakeApp", (), {})()
        global_vars.otl_wizard.main_window = main_window
        qtbot.addWidget(main_window)
        main_window.show()
        assert main_window.isVisible()

        # Arrange: Patch file picker for subset
        patch_subset_file_picker(slagboom_subset)

        # Arrange: Create and open project, get class table
        class_table = await create_and_open_project(qtbot, main_window, project_name, "TestBestek", slagboom_subset)
        qtbot.waitUntil(lambda: class_table.count() > 0, timeout=10000)

        # Assert: all classes initially selected
        selected_classes = [class_table.item(r).text() for r in range(class_table.count()) if class_table.item(r).isSelected()]
        all_classes = [class_table.item(r).text() for r in range(class_table.count())]
        assert selected_classes, "No classes are selected initially"
        assert set(selected_classes) == set(all_classes), "Not all classes are initially selected"

        # Act: Deselect Contactor and Kokerafsluiting
        select_classes(class_table, exclude={"Contactor", "Kokerafsluiting"})

        # Act: Update label and export button state
        main_window.step1.update_label_under_list(
            total_amount_of_items=class_table.count(),
            counter=len([i for i in range(class_table.count()) if class_table.item(i).isSelected()])
        )
        main_window.step1.export_button.setEnabled(True)

        # Act: Set all export options
        set_export_options(main_window.step1, choice_list=True, geometry=True, attr_info=True, example=True, expanded_info=True)

        # Act: Export Excel and wait for file
        export_btn = main_window.step1.export_button
        await export_template_and_wait_for_file(qtbot, export_btn, export_path)

        # Assert: Only the 4 classes are present in the Excel
        expected_classes = {
            'https://wegenenverkeer.data.vlaanderen.be/ns/installatie#Slagboom',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomarm',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#SlagboomarmVerlichting',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomkolom'
        }
        verify_exported_classes_in_excel(export_path, expected_classes)

        # Act: Remove first row from each sheet and save
        remove_first_row_from_excel(export_path)

    finally:
        # Clean up
        if project_dir.exists():
            shutil.rmtree(project_dir)
        if export_path.exists():
            export_path.unlink()


# --- Helper for CSV export and verification ---

def set_export_options_for_csv(step1):
    # Arrange: set export options for CSV (choice_list not available)
    step1.add_choice_list.setChecked(False)
    step1.add_geometry_attributes.setChecked(True)
    step1.export_attribute_info.setChecked(True)
    step1.example_amount_checkbox.setChecked(True)
    step1.radio_button_expanded_info.setChecked(True)


async def export_template_and_wait_for_file_csv(qtbot, export_btn, export_base_path, expected_classes, timeout=10000):
    """
    Clicks the export button and waits for all expected class CSV files to be created.

    Args:
        qtbot: The pytest-qt qtbot fixture.
        export_btn: The export button widget.
        export_base_path: The base path for the export files.
        expected_classes: List of expected class URIs.
        timeout: Timeout in milliseconds.

    Returns:
        List of Path objects for the created class CSV files.
    """
    def uri_to_filename(uri):
        ns, name = uri.split("#")
        ns_part = ns.rstrip("/").split("/")[-1]
        return f"{ns_part}_{name}"

    # Clean up any pre-existing files for all expected class files (only use correct pattern)
    for class_uri in expected_classes:
        class_file = export_base_path.parent / f"{export_base_path.stem}_{uri_to_filename(class_uri)}.csv"
        if class_file.exists():
            class_file.unlink()

    # Click the export button to trigger export
    qtbot.mouseClick(export_btn, Qt.MouseButton.LeftButton)

    class_files = [
        export_base_path.parent / f"{export_base_path.stem}_{uri_to_filename(uri)}.csv"
        for uri in expected_classes
    ]
    start = time.time()

    def file_ready(f):
        if not f.exists() or f.stat().st_size == 0:
            return False
        try:
            with open(f, newline='', encoding='utf-8') as csvfile:
                rows = list(csv.reader(csvfile))
                return len(rows) >= 3
        except Exception:
            return False

    while True:
        qtbot.wait(100)
        await asyncio.sleep(0)
        if all(file_ready(f) for f in class_files):
            break
        if (time.time() - start) * 1000 > timeout:
            raise TimeoutError(f"Not all class CSV files were created after {timeout} ms")
    for f in class_files:
        assert f.exists(), f"File {f} was not created"
    return class_files


def verify_exported_classes_in_csv(class_files, expected_classes):
    # Assert: check only expected classes are present in CSV
    found_classes = set()
    for class_file, class_uri in zip(class_files, expected_classes):
        with open(class_file, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            rows = list(reader)
            # Assume: first row is description, second row is headers, third row is data
            if len(rows) < 3:
                raise AssertionError(f"CSV file {class_file} does not have enough rows for data")
            # The class URI is in the first column of the data row (row 2, index 2)
            uri_value = rows[2][0].split(';')[0]
            assert uri_value == class_uri, f"Unexpected class URI in {class_file}: {uri_value}"
            found_classes.add(uri_value)
    missing = set(expected_classes) - found_classes
    assert not missing, f"Not all expected classes were found in the CSV. Missing: {missing}"


def remove_first_row_from_csv(class_files):
    # Act: remove first row (description) from each CSV and save
    for class_file in class_files:
        with open(class_file, newline='', encoding='utf-8') as csvfile:
            rows = list(csv.reader(csvfile))
        with open(class_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in rows[1:]:
                writer.writerow(row)

# --- Test 5: CSV export ---

@pytest.fixture(autouse=True)
def patch_helpers_open_folder_and_select_document(monkeypatch):
    """Monkeypatch Helpers.open_folder_and_select_document to do nothing in tests."""
    import otlmow_gui.Domain.util.Helpers as helpers_mod
    monkeypatch.setattr(helpers_mod.Helpers, "open_folder_and_select_document", staticmethod(lambda *args, **kwargs: None))


@pytest.mark.gui
@pytest.mark.asyncio
async def test_1_5_template_generation_and_class_selection_csv(qtbot, patch_subset_file_picker):
    """
    TEST stap 5: Genereer een CSV file met:
    - uitgebreide info
    - alle andere opties aan (behalve keuzelijsten(gaat niet))
    - check of alleen de aangeduide 4 klasses aanwezig zijn
    - Verwijder in elk bestand de eerste RIJ met omschrijvingen en save
    """
    project_name = "FullTestRunProject_1b_5"
    project_dir = Path.home() / 'OTLWizardProjects' / 'Projects' / project_name
    export_path = Path(tempfile.gettempdir()) / f"{project_name}_template.csv"
    slagboom_subset = Path(__file__).parent.parent.parent / "otlmow_gui" / "demo_projects" / "slagbomen_project" / "voorbeeld-slagboom.db"

    # Clean up before test
    if project_dir.exists():
        shutil.rmtree(project_dir)
    if export_path.exists():
        export_path.unlink()

    try:
        # Arrange: App and main window
        settings = Settings.get_or_create_settings_file()
        OTLLogger.init()
        language = GlobalTranslate(settings, LANG_DIR).get_all()
        main_window = MainWindow(language)
        global_vars.otl_wizard = type("FakeApp", (), {})()
        global_vars.otl_wizard.main_window = main_window
        qtbot.addWidget(main_window)
        main_window.show()
        assert main_window.isVisible()

        # Arrange: Patch file picker for subset
        patch_subset_file_picker(slagboom_subset)

        # Arrange: Create and open project, get class table
        class_table = await create_and_open_project(qtbot, main_window, project_name, "TestBestek", slagboom_subset)
        qtbot.waitUntil(lambda: class_table.count() > 0, timeout=10000)

        # Assert: all classes initially selected
        selected_classes = [class_table.item(r).text() for r in range(class_table.count()) if class_table.item(r).isSelected()]
        all_classes = [class_table.item(r).text() for r in range(class_table.count())]
        assert selected_classes, "No classes are selected initially"
        assert set(selected_classes) == set(all_classes), "Not all classes are initially selected"

        # Act: Deselect Contactor and Kokerafsluiting
        select_classes(class_table, exclude={"Contactor", "Kokerafsluiting"})

        # Act: Update label and export button state
        main_window.step1.update_label_under_list(
            total_amount_of_items=class_table.count(),
            counter=len([i for i in range(class_table.count()) if class_table.item(i).isSelected()])
        )
        main_window.step1.export_button.setEnabled(True)

        # Act: Select CSV as export file type
        main_window.step1.file_extension_selection.setCurrentText("CSV")
        qtbot.wait(100)  # Let the GUI update options

        # Act: Set export options for CSV
        set_export_options_for_csv(main_window.step1)

        # Act: Export CSV and wait for all class files
        export_btn = main_window.step1.export_button
        expected_classes = [
            'https://wegenenverkeer.data.vlaanderen.be/ns/installatie#Slagboom',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomarm',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#SlagboomarmVerlichting',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomkolom'
        ]
        class_files = await export_template_and_wait_for_file_csv(qtbot, export_btn, export_path, expected_classes)

        # Assert: Only the 4 classes are present in the CSVs
        verify_exported_classes_in_csv(class_files, expected_classes)

        # Act: Remove first row from each CSV and save
        remove_first_row_from_csv(class_files)
    finally:
        # Clean up
        if project_dir.exists():
            shutil.rmtree(project_dir)
        # Remove all class CSV files
        def uri_to_filename(uri):
            ns, name = uri.split("#")
            ns_part = ns.rstrip("/").split("/")[-1]
            return f"{ns_part}_{name}"

        for class_uri in [
            'https://wegenenverkeer.data.vlaanderen.be/ns/installatie#Slagboom',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomarm',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#SlagboomarmVerlichting',
            'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomkolom'
        ]:
            class_file = export_path.parent / f"{export_path.stem}_{uri_to_filename(class_uri)}.csv"
            if class_file.exists():
                class_file.unlink()