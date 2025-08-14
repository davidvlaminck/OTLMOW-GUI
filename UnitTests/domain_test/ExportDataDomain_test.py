from openpyxl.reader.excel import load_workbook
from pytestqt.qtbot import QtBot
from pytestqt.plugin import qtbot

from otlmow_gui.Domain.step_domain.ExportDataDomain import ExportDataDomain
from otlmow_gui.GUI import InsertDataScreen
from otlmow_gui.GUI import RelationChangeScreen

from UnitTests.general_fixtures.GUIFixtures import *
from UnitTests.general_fixtures.DomainFixtures import *

@fixture
def root_directory() -> Path:
    return Path(__file__).parent.parent.parent



@fixture
def mock_screen(qtbot: QtBot, create_translations) -> InsertDataScreen:
    insert_data_screen = InsertDataScreen(GlobalTranslate.instance.get_all())
    InsertDataDomain.get_screen = Mock(return_value=insert_data_screen)
    return insert_data_screen

@fixture
def mock_rel_screen(qtbot: QtBot, create_translations) -> RelationChangeScreen:
    relation_change_screen = RelationChangeScreen(GlobalTranslate.instance.get_all())

    original_fill_class_list = RelationChangeScreen.fill_object_list
    RelationChangeScreen.fill_object_list = Mock()

    original_get_screen = RelationChangeDomain.get_screen
    RelationChangeDomain.get_screen = Mock(return_value=relation_change_screen)

    yield relation_change_screen

    RelationChangeScreen.fill_object_list = original_fill_class_list
    RelationChangeDomain.get_screen = original_get_screen


@fixture
def mock_fill_possible_relations_list(mock_rel_screen: RelationChangeScreen):
    mock_rel_screen.fill_possible_relations_list = Mock()



@fixture
def get_export_path_with_cleanup(root_directory: Path) -> Path:
    export_path = (root_directory / "UnitTests" / "project_files_test" / "OTLWizardProjects" / "TestFiles" /
                   "testExport.xlsx")
    yield export_path

    if export_path.exists():
        os.remove(export_path)



@pytest.mark.asyncio
async def test_unedited_generate_files(root_directory: Path,
                                 setup_simpel_vergelijking_template5,
                                 mock_screen: InsertDataScreen,
                                mock_fill_possible_relations_list: RelationChangeScreen,
                                setup_test_project,
                                mock_step3_visuals,
                                 get_export_path_with_cleanup:Path,
                                 mock_save_validated_assets_function,
                                 mock_load_validated_assets):



    test_object_lists_file_path: list[str] = setup_simpel_vergelijking_template5

    InsertDataDomain.add_files_to_backend_list(test_object_lists_file_path)

    await InsertDataDomain.load_and_validate_documents()

    export_path = get_export_path_with_cleanup

    await ExportDataDomain.generate_files(export_path, False, False)

    assert export_path.exists()

    wb = load_workbook(export_path)

    expected_wb = load_workbook(Path(test_object_lists_file_path[0]))

    if 'Keuzelijsten' in expected_wb.sheetnames:
        expected_wb.remove(expected_wb['Keuzelijsten'])

    expected_sheet_titles = ['ins#Verkeersbordopstelling',
                             'ond#Funderingsmassief',
                             'ond#Pictogram',
                             'ond#Verkeersbordsteun',
                             'ond#Bevestiging',
                             'ond#HoortBij',
                             'ond#LigtOp']

    assert list(wb.sheetnames)  == expected_sheet_titles

    list_expected_values = [list(sh.iter_rows(0,4,0,26,True)) for sh in  wb.worksheets]
    list_values = [list(sh.iter_rows(0,4,0,26,True))for sh in expected_wb.worksheets]

    assert list_values == list_expected_values

@pytest.mark.asyncio
async def test_add_remove_generate_files(root_directory: Path,setup_simpel_vergelijking_template5,
                                mock_project_home_path,
                                 mock_screen: InsertDataScreen,
                                mock_fill_possible_relations_list: RelationChangeScreen,
                                setup_test_project,
                                mock_step3_visuals,
                                 get_export_path_with_cleanup:Path,
                                   mock_save_validated_assets_function,
                                   mock_load_validated_assets):

    test_object_lists_file_path: list[str] = setup_simpel_vergelijking_template5

    InsertDataDomain.add_files_to_backend_list(test_object_lists_file_path)

    await InsertDataDomain.load_and_validate_documents()



    await RelationChangeDomain.set_possible_relations(RelationChangeDomain.shown_objects[0])

    bron_id = 'dummy_hxOTHWe'
    target_id = 'dummy_a'
    index = 0
    RelationChangeDomain.add_possible_relation_to_existing_relations(bron_id,target_id,index)

    to_remove_index = -1
    to_remove_typeURI = 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Bevestiging'
    to_remove_source_id = 'dummy_Q'
    to_remove_target_id = 'dummy_bcjseEAj'
    for i, relation in enumerate(RelationChangeDomain.existing_relations):
        if (relation.bronAssetId.identificator == to_remove_source_id and
            relation.typeURI == to_remove_typeURI and
            relation.doelAssetId.identificator == to_remove_target_id):
            to_remove_index = i

    assert to_remove_index != -1

    RelationChangeDomain.remove_existing_relation(to_remove_index)

    export_path = get_export_path_with_cleanup

    await ExportDataDomain.generate_files(export_path, False, False,synchronous=True)

    assert export_path.exists()

    wb = load_workbook(export_path)

    ref_path_expected_workbook = (root_directory / "UnitTests" / "project_files_test" / "OTLWizardProjects" / "reference_files" / "ref_test_add_remove_generate_files.xlsx")
    expected_wb = load_workbook(Path(ref_path_expected_workbook))

    if 'Keuzelijsten' in expected_wb.sheetnames:
        expected_wb.remove(expected_wb['Keuzelijsten'])

    expected_sheet_titles = ['ins#Verkeersbordopstelling',
                             'ond#Funderingsmassief',
                             'ond#Pictogram',
                             'ond#Verkeersbordsteun',
                             'ond#Bevestiging',
                             'ond#HoortBij',
                             'ond#LigtOp']

    assert list(wb.sheetnames)  == expected_sheet_titles

    list_expected_values = [list(sh.iter_rows(0,4,0,26,True)) for sh in  wb.worksheets]
    list_values = [list(sh.iter_rows(0,4,0,26,True))for sh in expected_wb.worksheets]

    assert list_values == list_expected_values

@fixture
def get_export_path_export_with_cleanup(root_directory: Path) -> Path:
    export_path = (root_directory / "UnitTests" / "project_files_test" / "OTLWizardProjects" / "TestFiles" /
                   "testExport.xlsx")
    yield export_path

    if export_path.exists():
        os.remove(export_path)

@pytest.mark.asyncio
async def test_add_remove_inactive_relations_and_generate_files(root_directory: Path,
                                                          setup_simpel_vergelijking_template5,
                                 mock_screen: InsertDataScreen,
                                mock_fill_possible_relations_list: RelationChangeScreen,
                                setup_test_project,
                                mock_step3_visuals,
                                 get_export_path_export_with_cleanup:Path,
                                   mock_save_validated_assets_function,
                                   mock_load_validated_assets):

    test_object_lists_file_path: list[str] = setup_simpel_vergelijking_template5

    InsertDataDomain.add_files_to_backend_list(test_object_lists_file_path)

    await InsertDataDomain.load_and_validate_documents()

    source_id_1  = 'dummy_FNrHuPZCWV'
    target_id_1 = 'dummy_TyBGmXfXC'
    index_1 = 1

    await RelationChangeDomain.set_possible_relations( RelationChangeDomain.get_object(source_id_1))
    added_relation =  RelationChangeDomain.add_possible_relation_to_existing_relations(source_id_1,
                                                                     target_id_1,
                                                                     index_1)
    print(f"first added relation {added_relation.typeURI}")

    source_id_2 = 'dummy_a'
    target_id_2 = 'dummy_TyBGmXfXC'
    index_2 = 0

    await RelationChangeDomain.set_possible_relations(
        selected_object=RelationChangeDomain.get_object(identificator=source_id_2))
    added_relation =  RelationChangeDomain.add_possible_relation_to_existing_relations(
        bron_asset_id=source_id_2,
        target_asset_id=target_id_2,
        relation_object_index= index_2)
    print(f"second added relation {added_relation.typeURI}" )

    to_remove_index = -1
    to_remove_typeURI = 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Bevestiging'
    to_remove_source_id = 'dummy_Q'
    to_remove_target_id = 'dummy_bcjseEAj'
    for i, relation in enumerate(RelationChangeDomain.existing_relations):
        if (relation.bronAssetId.identificator == to_remove_source_id and
                relation.typeURI == to_remove_typeURI and
                relation.doelAssetId.identificator == to_remove_target_id):
            to_remove_index = i

    assert to_remove_index != -1

    remove_relation =  RelationChangeDomain.remove_existing_relation(to_remove_index)
    print(f"removed relation bron id {remove_relation.bronAssetId.identificator} target id {remove_relation.doelAssetId.identificator}  type {remove_relation.typeURI}")

    export_path = get_export_path_export_with_cleanup

    await ExportDataDomain.generate_files(export_path, False, False,synchronous=True)

    assert export_path.exists()

    wb = load_workbook(export_path)

    ref_path_expected_workbook = (root_directory / "UnitTests" / "project_files_test" /
                                  "OTLWizardProjects" / "reference_files" /
                                  "ref_test_add_temove_inactive_relations_and_generate_files.xlsx")
    expected_wb = load_workbook(Path(ref_path_expected_workbook))

    if 'Keuzelijsten' in expected_wb.sheetnames:
        expected_wb.remove(expected_wb['Keuzelijsten'])

    expected_sheet_titles = ['ins#Verkeersbordopstelling',
                             'ond#Funderingsmassief',
                             'ond#Pictogram',
                             'ond#Verkeersbordsteun',
                             'ond#Bevestiging',
                             'ond#HoortBij',
                             'ond#LigtOp']

    assert list(wb.sheetnames)  == expected_sheet_titles

    list_expected_values = [list(sh.iter_rows(0,4,0,26,True)) for sh in  wb.worksheets]
    list_values = [list(sh.iter_rows(0,4,0,26,True))for sh in expected_wb.worksheets]

    assert list_values == list_expected_values